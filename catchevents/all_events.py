#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2021/9/6 7:18 下午
# @Author  : Fu Yiyi
# @Site    : 
# @File    : all_events.py
# @Software: PyCharm

import openpyxl


def get_all_events(file):
    if file.__contains__("xlsx"):
        # 打开Excel文件读取数据
        workSheet = openpyxl.load_workbook(file)
        # 获取一个工作表
        sheetName = workSheet.sheetnames[0]
        sheetName = workSheet[sheetName]
        # 获取行列数量
        row = len([row for row in sheetName if not all([cell.value is None for cell in row])])
        # col = sheetName.max_column
        events = list()
        for i in range(1, row + 1):
            if sheetName.cell(i, 1).value is None:
                continue
            event = sheetName.cell(i, 2).value
            events.append(event)
            # events_num = len(events)
        return events
